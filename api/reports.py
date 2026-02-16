"""
Reports API endpoints
"""
from flask import Blueprint, request, jsonify, session, send_file
from flask_login import login_required
from models import db, Item, SaleItem, PurchaseItem, Sale, PurchaseContainer, Company, SafeTransaction, SafeStatementRealBalance, Market, InventoryAdjustment, InventoryBatch, SaleItemAllocation, Payment, GeneralExpense
from decimal import Decimal
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from sqlalchemy import func, case
from sqlalchemy.orm import joinedload

bp = Blueprint('reports', __name__)

@bp.route('/daily-sales', methods=['GET'])
@login_required
def get_daily_sales():
    """Get daily sales grouped by date, with all sales for each day combined into one invoice"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Build query
    query = Sale.query.filter_by(market_id=market_id)
    
    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    sales = query.order_by(Sale.date.asc(), Sale.id.asc()).all()
    
    # Group sales by date
    daily_sales = {}
    for sale in sales:
        sale_date = sale.date.isoformat()
        
        if sale_date not in daily_sales:
            daily_sales[sale_date] = {
                'date': sale_date,
                'sales': [],
                'total_amount': Decimal('0'),
                'total_paid': Decimal('0'),
                'total_balance': Decimal('0'),
                'customers': set(),
                'suppliers': set()
            }
        
        # Get sale items
        items = [{
            'id': i.id,
            'item_id': i.item_id,
            'item_code': i.item.code,
            'item_name': i.item.name,
            'quantity': float(i.quantity),
            'unit_price': float(i.unit_price),
            'total_price': float(i.total_price),
            'customer_name': sale.customer.name,
            'supplier_name': sale.supplier.name if sale.supplier else None
        } for i in sale.items]
        
        # Get supplier name
        supplier_name = None
        if sale.supplier_id:
            supplier = Company.query.get(sale.supplier_id)
            supplier_name = supplier.name if supplier else None
        
        daily_sales[sale_date]['sales'].append({
            'id': sale.id,
            'invoice_number': sale.invoice_number,
            'customer_id': sale.customer_id,
            'customer_name': sale.customer.name,
            'supplier_id': sale.supplier_id,
            'supplier_name': supplier_name,
            'total_amount': float(sale.total_amount),
            'paid_amount': float(sale.paid_amount),
            'balance': float(sale.balance),
            'payment_type': sale.payment_type,
            'status': sale.status,
            'items': items
        })
        
        daily_sales[sale_date]['total_amount'] += sale.total_amount
        daily_sales[sale_date]['total_paid'] += sale.paid_amount
        daily_sales[sale_date]['total_balance'] += sale.balance
        daily_sales[sale_date]['customers'].add(sale.customer.name)
        if supplier_name:
            daily_sales[sale_date]['suppliers'].add(supplier_name)
    
    # Convert to list format
    result = []
    for date, data in sorted(daily_sales.items()):
        result.append({
            'date': data['date'],
            'total_amount': float(data['total_amount']),
            'total_paid': float(data['total_paid']),
            'total_balance': float(data['total_balance']),
            'customers': list(data['customers']),
            'suppliers': list(data['suppliers']),
            'sales': data['sales']
        })
    
    return jsonify(result)

@bp.route('/safe-statement', methods=['GET'])
@login_required
def get_safe_statement():
    """Get safe statement with daily totals (IN/OUT) and real balance"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    export_excel = request.args.get('export') == 'excel'
    
    # Get all transactions in date range
    query = SafeTransaction.query.filter_by(market_id=market_id)
    
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(SafeTransaction.date >= start_date_obj)
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(SafeTransaction.date <= end_date_obj)
    
    transactions = query.order_by(SafeTransaction.date.asc(), SafeTransaction.id.asc()).all()
    
    # Group by date
    daily_totals = {}
    for txn in transactions:
        date_str = txn.date.isoformat()
        if date_str not in daily_totals:
            daily_totals[date_str] = {
                'date': date_str,
                'total_in': Decimal('0'),
                'total_out': Decimal('0'),
                'balance': Decimal('0'),
                'real_balance': None
            }
        
        if txn.transaction_type in ['Opening', 'Inflow']:
            daily_totals[date_str]['total_in'] += txn.amount_base_currency
        elif txn.transaction_type == 'Outflow':
            daily_totals[date_str]['total_out'] += txn.amount_base_currency
        
        # Balance is the balance_after from the last transaction of the day
        daily_totals[date_str]['balance'] = txn.balance_after
    
    # Get real balances from SafeStatementRealBalance
    if daily_totals:
        date_list = list(daily_totals.keys())
        real_balances = SafeStatementRealBalance.query.filter_by(
            market_id=market_id
        ).filter(
            SafeStatementRealBalance.date.in_([datetime.strptime(d, '%Y-%m-%d').date() for d in date_list])
        ).all()
        
        for rb in real_balances:
            date_str = rb.date.isoformat()
            if date_str in daily_totals:
                daily_totals[date_str]['real_balance'] = float(rb.real_balance) if rb.real_balance else None
    
    # Convert to list and sort by date
    statement = []
    for date_str in sorted(daily_totals.keys()):
        data = daily_totals[date_str]
        statement.append({
            'date': data['date'],
            'total_in': float(data['total_in']),
            'total_out': float(data['total_out']),
            'balance': float(data['balance']),
            'real_balance': data['real_balance']
        })
    
    # Export to Excel if requested
    if export_excel:
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = pd.DataFrame(statement)
            df.to_excel(writer, index=False, sheet_name='Safe Statement')
            
            # Format columns
            worksheet = writer.sheets['Safe Statement']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
        
        output.seek(0)
        filename = f'safe_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                         as_attachment=True, download_name=filename)
    
    return jsonify({'statement': statement})

@bp.route('/safe-statement/real-balance', methods=['GET', 'PUT'])
@login_required
def safe_statement_real_balance():
    """Get or update real balance for a specific date"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    if request.method == 'GET':
        date_str = request.args.get('date')
        if not date_str:
            return jsonify({'error': 'Date is required'}), 400
        
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        real_balance = SafeStatementRealBalance.query.filter_by(
            market_id=market_id,
            date=date_obj
        ).first()
        
        return jsonify({
            'date': date_str,
            'real_balance': float(real_balance.real_balance) if real_balance and real_balance.real_balance else None
        })
    
    elif request.method == 'PUT':
        data = request.json
        date_str = data.get('date')
        real_balance_value = data.get('real_balance')
        
        if not date_str:
            return jsonify({'error': 'Date is required'}), 400
        
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Find or create real balance record
        real_balance = SafeStatementRealBalance.query.filter_by(
            market_id=market_id,
            date=date_obj
        ).first()
        
        if not real_balance:
            real_balance = SafeStatementRealBalance(
                market_id=market_id,
                date=date_obj,
                real_balance=Decimal(str(real_balance_value)) if real_balance_value is not None else None
            )
            db.session.add(real_balance)
        else:
            real_balance.real_balance = Decimal(str(real_balance_value)) if real_balance_value is not None else None
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'date': date_str,
            'real_balance': float(real_balance.real_balance) if real_balance.real_balance else None
        })

# Placeholder endpoints to prevent 404 errors
# These return minimal data structures expected by the frontend

@bp.route('/profit-loss', methods=['GET'])
@login_required
def get_profit_loss():
    """Get profit & loss report"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    market = Market.query.get(market_id)
    calculation_method = getattr(market, 'calculation_method', 'Average') if market else 'Average'
    base_currency = market.base_currency if market else 'USD'
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    item_id = request.args.get('item_id', type=int)
    
    # Build query - eager load item and sale to avoid N+1
    query = SaleItem.query.options(
        joinedload(SaleItem.item),
        joinedload(SaleItem.sale)
    ).join(Sale).filter(Sale.market_id == market_id)
    
    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if item_id:
        query = query.filter(SaleItem.item_id == item_id)
    
    sale_items = query.all()
    
    if not sale_items:
        return jsonify({
            'calculation_method': calculation_method,
            'base_currency': base_currency,
            'items': [],
            'totals': {'total_sales': 0, 'total_cog': 0, 'total_cost': 0, 'total_profit': 0, 'profit_margin': 0}
        })
    
    sale_item_ids = [si.id for si in sale_items]
    unique_item_ids = list(set(si.item_id for si in sale_items))
    
    # Pre-load all data in bulk to avoid N+1 queries
    if calculation_method == 'FIFO':
        allocations = SaleItemAllocation.query.filter(
            SaleItemAllocation.sale_item_id.in_(sale_item_ids)
        ).all()
        batch_ids = list(set(a.batch_id for a in allocations))
        batches = {b.id: b for b in InventoryBatch.query.filter(InventoryBatch.id.in_(batch_ids)).all()} if batch_ids else {}
        allocations_by_sale_item = {}
        for a in allocations:
            allocations_by_sale_item.setdefault(a.sale_item_id, []).append(a)
    else:
        # Average: pre-load purchase items for relevant item_ids, then all items per container for COG
        purchase_items_all = PurchaseItem.query.options(
            joinedload(PurchaseItem.item),
            joinedload(PurchaseItem.container)
        ).join(PurchaseContainer).filter(
            PurchaseContainer.market_id == market_id,
            PurchaseItem.item_id.in_(unique_item_ids)
        ).all()
        container_ids = list(set(pi.container_id for pi in purchase_items_all))
        # Load ALL items per container (needed for COG: total_qty and total_weight of entire container)
        all_container_items = PurchaseItem.query.options(joinedload(PurchaseItem.item)).filter(
            PurchaseItem.container_id.in_(container_ids)
        ).all() if container_ids else []
        container_items_map = {}
        for pi in all_container_items:
            container_items_map.setdefault(pi.container_id, []).append(pi)
        
        avg_cost_per_item = {}
        avg_cost_supplier_per_item = {}
        purchase_by_item = {}
        for pi in purchase_items_all:
            purchase_by_item.setdefault(pi.item_id, []).append(pi)
        
        for iid in unique_item_ids:
            purchase_items = purchase_by_item.get(iid, [])
            if not purchase_items:
                continue
            total_cost_base = Decimal('0')
            total_cost_supplier = Decimal('0')
            total_qty = Decimal('0')
            supplier_currency = None
            for pi in purchase_items:
                container = pi.container
                if not container:
                    continue
                container_id = pi.container_id
                if supplier_currency is None:
                    supplier_currency = container.currency
                all_items = container_items_map.get(container_id, [])
                total_qty_container = sum(p.quantity for p in all_items)
                total_weight_container = sum((p.item.weight or Decimal('0')) * p.quantity for p in all_items)
                expense1 = Decimal('0')
                if container.expense1_amount and container.expense1_amount > 0:
                    if container.expense1_currency == container.currency:
                        expense1 = container.expense1_amount
                    else:
                        eb = container.expense1_amount * (container.expense1_exchange_rate or 1)
                        cr = container.exchange_rate or 1
                        if cr > 0:
                            expense1 = eb / cr
                expense2 = Decimal('0')
                if container.expense2_amount and container.expense2_amount > 0:
                    if container.expense2_currency == container.currency:
                        expense2 = container.expense2_amount
                    else:
                        eb = container.expense2_amount * (container.expense2_exchange_rate or 1)
                        cr = container.exchange_rate or 1
                        if cr > 0:
                            expense2 = eb / cr
                expense3 = Decimal('0')
                if container.expense3_amount and container.expense3_amount > 0:
                    if container.expense3_currency == container.currency:
                        expense3 = container.expense3_amount
                    else:
                        eb = container.expense3_amount * (container.expense3_exchange_rate or 1)
                        cr = container.exchange_rate or 1
                        if cr > 0:
                            expense3 = eb / cr
                sum_expenses = expense1 + expense2 + expense3
                item_weight = pi.item.weight or Decimal('0')
                if total_qty_container > 0 and total_weight_container > 0:
                    cog_per_unit = (sum_expenses / Decimal('2') / total_qty_container) + \
                                  (sum_expenses / Decimal('2') / total_weight_container * item_weight)
                elif total_qty_container > 0:
                    cog_per_unit = sum_expenses / total_qty_container
                else:
                    cog_per_unit = Decimal('0')
                item_cost_per_unit_supplier = pi.unit_price + cog_per_unit
                item_cost_per_unit_base = item_cost_per_unit_supplier * container.exchange_rate
                total_cost_base += item_cost_per_unit_base * pi.quantity
                total_cost_supplier += item_cost_per_unit_supplier * pi.quantity
                total_qty += pi.quantity
            if total_qty > 0:
                avg_cost_per_item[iid] = total_cost_base / total_qty
                avg_cost_supplier_per_item[iid] = (total_cost_supplier / total_qty, supplier_currency)
    
    # Group by item (single pass, no queries)
    items_data = {}
    for si in sale_items:
        iid = si.item_id
        if iid not in items_data:
            items_data[iid] = {
                'item_id': iid,
                'item_code': si.item.code,
                'item_name': si.item.name,
                'quantity_sold': Decimal('0'),
                'total_sales': Decimal('0'),
                'total_cost': Decimal('0'),
                'cog': Decimal('0'),
                'batch_details': []
            }
        
        items_data[iid]['quantity_sold'] += si.quantity
        items_data[iid]['total_sales'] += si.total_price
        
        if calculation_method == 'FIFO':
            for alloc in allocations_by_sale_item.get(si.id, []):
                batch = batches.get(alloc.batch_id)
                if batch:
                    cost_per_unit_base = batch.cost_per_unit * (batch.exchange_rate or Decimal('1'))
                    item_cost = cost_per_unit_base * alloc.quantity
                    items_data[iid]['total_cost'] += item_cost
                    items_data[iid]['cog'] += item_cost
                    items_data[iid]['batch_details'].append({
                        'sale_date': si.sale.date.isoformat(),
                        'invoice_number': si.sale.invoice_number,
                        'batch_code': batch.container.container_number if batch.container else '',
                        'purchase_date': batch.purchase_date.isoformat() if batch.purchase_date else None,
                        'quantity': float(alloc.quantity),
                        'cost_per_unit': float(cost_per_unit_base),
                        'total_cost': float(item_cost),
                        'currency': base_currency
                    })
        else:
            avg_cost = avg_cost_per_item.get(iid)
            if avg_cost is not None:
                item_cost = avg_cost * si.quantity
                items_data[iid]['total_cost'] += item_cost
                items_data[iid]['cog'] += item_cost
            avg_supplier_info = avg_cost_supplier_per_item.get(iid)
            if avg_supplier_info is not None:
                items_data[iid]['average_purchase_price_supplier_currency'] = float(avg_supplier_info[0])
                items_data[iid]['supplier_currency'] = avg_supplier_info[1]
    
    # Convert to list and calculate profit
    items_list = []
    total_sales = Decimal('0')
    total_cog = Decimal('0')
    total_cost = Decimal('0')
    total_profit = Decimal('0')
    
    for item_data in items_data.values():
        profit = item_data['total_sales'] - item_data['total_cost']
        profit_margin = (profit / item_data['total_sales'] * 100) if item_data['total_sales'] > 0 else 0
        
        row = {
            'item_code': item_data['item_code'],
            'item_name': item_data['item_name'],
            'quantity_sold': float(item_data['quantity_sold']),
            'total_sales': float(item_data['total_sales']),
            'cog': float(item_data['cog']),
            'average_purchase_price': float(item_data['total_cost'] / item_data['quantity_sold']) if item_data['quantity_sold'] > 0 else 0,
            'total_cost': float(item_data['total_cost']),
            'profit': float(profit),
            'profit_margin': float(profit_margin),
            'batch_details': item_data['batch_details'] if calculation_method == 'FIFO' else []
        }
        if calculation_method != 'FIFO':
            row['average_purchase_price_supplier_currency'] = item_data.get('average_purchase_price_supplier_currency')
            row['supplier_currency'] = item_data.get('supplier_currency')
        items_list.append(row)
        
        total_sales += item_data['total_sales']
        total_cog += item_data['cog']
        total_cost += item_data['total_cost']
        total_profit += profit
    
    profit_margin = (total_profit / total_sales * 100) if total_sales > 0 else 0
    
    return jsonify({
        'calculation_method': calculation_method,
        'base_currency': base_currency,
        'items': items_list,
        'totals': {
            'total_sales': float(total_sales),
            'total_cog': float(total_cog),
            'total_cost': float(total_cost),
            'total_profit': float(total_profit),
            'profit_margin': float(profit_margin)
        }
    })

@bp.route('/customer-receivables', methods=['GET'])
@login_required
def get_customer_receivables():
    """Get customer receivables report"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    customers = Company.query.filter_by(market_id=market_id, category='Customer').all()
    receivables = []
    total_receivables = Decimal('0')
    
    for customer in customers:
        balance = customer.get_balance(market_id)
        if balance > 0:
            receivables.append({
                'customer_id': customer.id,
                'customer_name': customer.name,
                'currency': customer.currency,
                'balance': float(balance)
            })
            total_receivables += Decimal(str(balance))
    
    return jsonify({
        'receivables': receivables,
        'total_receivables': float(total_receivables)
    })

@bp.route('/supplier-payables', methods=['GET'])
@login_required
def get_supplier_payables():
    """Get supplier payables report"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    suppliers = Company.query.filter_by(market_id=market_id, category='Supplier').all()
    payables = []
    currency_totals = {}
    
    for supplier in suppliers:
        balance = supplier.get_balance(market_id)
        if balance > 0:
            currency = supplier.currency
            payables.append({
                'supplier_id': supplier.id,
                'supplier_name': supplier.name,
                'currency': currency,
                'balance': float(balance)
            })
            
            if currency not in currency_totals:
                currency_totals[currency] = Decimal('0')
            currency_totals[currency] += Decimal(str(balance))
    
    return jsonify({
        'payables': payables,
        'currency_totals': {k: float(v) for k, v in currency_totals.items()}
    })

@bp.route('/sales', methods=['GET'])
@login_required
def get_sales_report():
    """Get sales report by item"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    market = Market.query.get(market_id)
    base_currency = market.base_currency if market else 'USD'
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = SaleItem.query.join(Sale).filter(Sale.market_id == market_id)
    
    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    sale_items = query.all()
    
    # Group by item
    items_data = {}
    for si in sale_items:
        item_id = si.item_id
        if item_id not in items_data:
            items_data[item_id] = {
                'item_id': item_id,
                'item_code': si.item.code,
                'item_name': si.item.name,
                'total_quantity': Decimal('0'),
                'total_amount': Decimal('0'),
                'sales': []
            }
        
        items_data[item_id]['total_quantity'] += si.quantity
        items_data[item_id]['total_amount'] += si.total_price
        
        supplier_name = si.sale.supplier.name if si.sale.supplier else None
        
        items_data[item_id]['sales'].append({
            'date': si.sale.date.isoformat(),
            'invoice_number': si.sale.invoice_number,
            'customer_name': si.sale.customer.name if si.sale.customer else 'Unknown',
            'supplier_name': supplier_name,
            'quantity': float(si.quantity),
            'unit_price': float(si.unit_price),
            'total_price': float(si.total_price),
            'payment_type': si.sale.payment_type,
            'status': si.sale.status
        })
    
    items_list = []
    total_sales = Decimal('0')
    total_items_sold = Decimal('0')
    transactions_count = 0
    
    for item_data in items_data.values():
        items_list.append({
            'item_code': item_data['item_code'],
            'item_name': item_data['item_name'],
            'total_quantity': float(item_data['total_quantity']),
            'total_amount': float(item_data['total_amount']),
            'sales': item_data['sales']
        })
        total_sales += item_data['total_amount']
        total_items_sold += item_data['total_quantity']
        transactions_count += len(item_data['sales'])
    
    return jsonify({
        'base_currency': base_currency,
        'items': items_list,
        'totals': {
            'total_sales': float(total_sales),
            'total_items_sold': float(total_items_sold),
            'items_count': len(items_list),
            'transactions_count': transactions_count
        }
    })

@bp.route('/sales/export', methods=['GET'])
@login_required
def export_sales_report():
    """Export sales report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    market = Market.query.get(market_id)
    base_currency = market.base_currency if market else 'USD'
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = SaleItem.query.join(Sale).filter(Sale.market_id == market_id)
    
    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    sale_items = query.order_by(Sale.date.asc(), Sale.id.asc()).all()
    
    # Build export data - flatten the structure for Excel
    export_data = []
    
    for si in sale_items:
        supplier_name = si.sale.supplier.name if si.sale.supplier else None
        export_data.append({
            'Date': si.sale.date.isoformat(),
            'Invoice Number': si.sale.invoice_number,
            'Item Code': si.item.code,
            'Item Name': si.item.name,
            'Customer': si.sale.customer.name if si.sale.customer else 'Unknown',
            'Supplier': supplier_name or '',
            'Quantity': float(si.quantity),
            'Unit Price': float(si.unit_price),
            'Total Price': float(si.total_price),
            'Payment Type': si.sale.payment_type,
            'Status': si.sale.status
        })
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Sales Report')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Sales Report']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
    
    output.seek(0)
    filename = f'sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# Add placeholder endpoints for other missing reports
@bp.route('/inventory-stock', methods=['GET'])
@login_required
def get_inventory_stock():
    """Get inventory stock report"""
    return jsonify({'error': 'Endpoint not yet implemented'}), 501

@bp.route('/inventory-snapshot', methods=['GET'])
@login_required
def get_inventory_snapshot():
    """Get inventory snapshot report"""
    return jsonify({'error': 'Endpoint not yet implemented'}), 501

@bp.route('/container-report', methods=['GET'])
@login_required
def get_container_report():
    """Get container report with items, expenses, and costs"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    container_id = request.args.get('container_id', type=int)
    if not container_id:
        return jsonify({'error': 'Container ID is required'}), 400
    
    # Get container
    container = PurchaseContainer.query.filter_by(
        market_id=market_id,
        id=container_id
    ).first()
    
    if not container:
        return jsonify({'error': 'Container not found'}), 404
    
    # Get all purchase items for this container
    purchase_items = PurchaseItem.query.filter_by(container_id=container.id).all()
    
    # Calculate expenses in container currency
    expense1 = Decimal('0')
    if container.expense1_amount and container.expense1_amount > 0:
        if container.expense1_currency == container.currency:
            expense1 = container.expense1_amount
        else:
            # Convert to base currency first, then to container currency
            expense1_base = container.expense1_amount * (container.expense1_exchange_rate or 1)
            container_rate = container.exchange_rate or 1
            if container_rate > 0:
                expense1 = expense1_base / container_rate
    
    expense2 = Decimal('0')
    if container.expense2_amount and container.expense2_amount > 0:
        if container.expense2_currency == container.currency:
            expense2 = container.expense2_amount
        else:
            # Convert to base currency first, then to container currency
            expense2_base = container.expense2_amount * (container.expense2_exchange_rate or 1)
            container_rate = container.exchange_rate or 1
            if container_rate > 0:
                expense2 = expense2_base / container_rate
    
    expense3 = Decimal('0')
    if container.expense3_amount and container.expense3_amount > 0:
        if container.expense3_currency == container.currency:
            expense3 = container.expense3_amount
        else:
            # Convert to base currency first, then to container currency
            expense3_base = container.expense3_amount * (container.expense3_exchange_rate or 1)
            container_rate = container.exchange_rate or 1
            if container_rate > 0:
                expense3 = expense3_base / container_rate
    
    sum_expenses = expense1 + expense2 + expense3
    
    # Calculate total quantity and total weight for the container
    total_quantity = sum(pi.quantity for pi in purchase_items)
    total_weight = sum((pi.item.weight or Decimal('0')) * pi.quantity for pi in purchase_items)
    
    # Build items list with all calculations
    items = []
    total_price = Decimal('0')
    total_cog = Decimal('0')
    total_item_cost = Decimal('0')
    
    for pi in purchase_items:
        item = pi.item
        item_weight = (item.weight or Decimal('0'))
        item_total_weight = item_weight * pi.quantity
        
        # Calculate COG per unit in container currency
        if total_quantity > 0 and total_weight > 0:
            cog_per_unit = (sum_expenses / Decimal('2') / total_quantity) + \
                          (sum_expenses / Decimal('2') / total_weight * item_weight)
        elif total_quantity > 0:
            # If no weight, distribute by quantity only
            cog_per_unit = sum_expenses / total_quantity
        else:
            cog_per_unit = Decimal('0')
        
        total_cog_for_item = cog_per_unit * pi.quantity
        
        # Item cost per unit = unit_price (in container currency) + COG per unit
        item_cost_per_unit = pi.unit_price + cog_per_unit
        item_total_cost_for_item = item_cost_per_unit * pi.quantity
        
        items.append({
            'item_code': item.code,
            'item_name': item.name,
            'quantity': float(pi.quantity),
            'item_weight': float(item_weight),
            'item_total_weight': float(item_total_weight),
            'unit_price': float(pi.unit_price),
            'total_price': float(pi.total_price),
            'cog': float(cog_per_unit),
            'total_cog': float(total_cog_for_item),
            'item_cost': float(item_cost_per_unit),
            'item_total_cost': float(item_total_cost_for_item)
        })
        
        total_price += pi.total_price
        total_cog += total_cog_for_item
        total_item_cost += item_total_cost_for_item
    
    # Calculate supplier cost (total price + expense1)
    # Note: total_item_cost already includes COG, so we use total_price + expense1
    # to avoid double-counting expense1
    supplier_cost = total_price + expense1
    
    return jsonify({
        'container': {
            'container_number': container.container_number,
            'supplier_name': container.supplier.name if container.supplier else 'Unknown',
            'currency': container.currency,
            'exchange_rate': float(container.exchange_rate),
            'date': container.date.isoformat()
        },
        'items': items,
        'totals': {
            'quantity': float(total_quantity),
            'item_total_weight': float(total_weight),
            'total_price': float(total_price),
            'total_cog': float(total_cog),
            'item_total_cost': float(total_item_cost)
        },
        'expenses': {
            'expense1': float(expense1),
            'expense2': float(expense2),
            'expense3': float(expense3),
            'total': float(sum_expenses)
        },
        'supplier_cost': float(supplier_cost)
    })

@bp.route('/stock-value-details', methods=['GET'])
@login_required
def get_stock_value_details():
    """Get stock value details report grouped by supplier"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    item_id = request.args.get('item_id', type=int)
    
    # Get market calculation method
    market = Market.query.get(market_id)
    calculation_method = getattr(market, 'calculation_method', 'Average') if market else 'Average'
    
    # Get all suppliers
    suppliers = Company.query.filter_by(market_id=market_id, category='Supplier').all()
    
    suppliers_data = []
    
    for supplier in suppliers:
        # Get items for this supplier
        items_query = Item.query.filter_by(market_id=market_id, supplier_id=supplier.id)
        if item_id:
            items_query = items_query.filter_by(id=item_id)
        items = items_query.all()
        
        if not items:
            continue
        
        supplier_items = []
        
        for item in items:
            # Get purchased quantity
            purchased_qty = db.session.query(func.coalesce(func.sum(PurchaseItem.quantity), 0)).join(
                PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
            ).filter(
                PurchaseContainer.market_id == market_id,
                PurchaseItem.item_id == item.id
            ).scalar() or Decimal('0')
            
            # Get sold quantity
            sold_qty = db.session.query(func.coalesce(func.sum(SaleItem.quantity), 0)).join(
                Sale, SaleItem.sale_id == Sale.id
            ).filter(
                Sale.market_id == market_id,
                SaleItem.item_id == item.id
            ).scalar() or Decimal('0')
            
            # Get inventory adjustments
            adjustment_qty = db.session.query(func.sum(
                case(
                    (InventoryAdjustment.adjustment_type == 'Increase', InventoryAdjustment.quantity),
                    else_=-InventoryAdjustment.quantity
                )
            )).filter(
                InventoryAdjustment.market_id == market_id,
                InventoryAdjustment.item_id == item.id
            ).scalar() or Decimal('0')
            
            available_qty = purchased_qty - sold_qty + adjustment_qty
            
            # Get containers for this item
            purchase_items = PurchaseItem.query.join(
                PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
            ).filter(
                PurchaseContainer.market_id == market_id,
                PurchaseItem.item_id == item.id
            ).all()
            
            containers_data = []
            total_cost_all_containers = Decimal('0')
            total_quantity_all_containers = Decimal('0')
            
            # Group by container - get unique container IDs
            container_ids = set(pi.container_id for pi in purchase_items)
            
            for container_id in container_ids:
                container = PurchaseContainer.query.get(container_id)
                if not container:
                    continue
                
                # Get ALL purchase items for this container (not just this item) for COG calculation
                all_container_items = PurchaseItem.query.filter_by(container_id=container_id).all()
                
                # Find the purchase item for this specific item
                item_pi = next((pi for pi in purchase_items if pi.container_id == container_id), None)
                if not item_pi:
                    continue
                
                # Calculate expenses in container currency
                expense1_in_container_currency = Decimal('0')
                expense1_original = Decimal('0')
                expense1_currency = container.currency
                if container.expense1_amount and container.expense1_amount > 0:
                    expense1_original = container.expense1_amount
                    expense1_currency = container.expense1_currency
                    if container.expense1_currency == container.currency:
                        expense1_in_container_currency = container.expense1_amount
                    else:
                        expense1_base = container.expense1_amount * (container.expense1_exchange_rate or 1)
                        container_rate = container.exchange_rate or 1
                        if container_rate > 0:
                            expense1_in_container_currency = expense1_base / container_rate
                
                expense2_in_container_currency = Decimal('0')
                expense2_original = Decimal('0')
                expense2_currency = container.currency
                if container.expense2_amount and container.expense2_amount > 0:
                    expense2_original = container.expense2_amount
                    expense2_currency = container.expense2_currency
                    if container.expense2_currency == container.currency:
                        expense2_in_container_currency = container.expense2_amount
                    else:
                        expense2_base = container.expense2_amount * (container.expense2_exchange_rate or 1)
                        container_rate = container.exchange_rate or 1
                        if container_rate > 0:
                            expense2_in_container_currency = expense2_base / container_rate
                
                expense3_in_container_currency = Decimal('0')
                expense3_original = Decimal('0')
                expense3_currency = container.currency
                if container.expense3_amount and container.expense3_amount > 0:
                    expense3_original = container.expense3_amount
                    expense3_currency = container.expense3_currency
                    if container.expense3_currency == container.currency:
                        expense3_in_container_currency = container.expense3_amount
                    else:
                        expense3_base = container.expense3_amount * (container.expense3_exchange_rate or 1)
                        container_rate = container.exchange_rate or 1
                        if container_rate > 0:
                            expense3_in_container_currency = expense3_base / container_rate
                
                sum_expenses = expense1_in_container_currency + expense2_in_container_currency + expense3_in_container_currency
                
                # Calculate total quantity and weight for ENTIRE container (all items)
                total_qty_container = sum(pi.quantity for pi in all_container_items)
                total_weight_container = sum((pi.item.weight or Decimal('0')) * pi.quantity for pi in all_container_items)
                
                item_weight = item.weight or Decimal('0')
                
                # Calculate COG per unit
                if total_qty_container > 0 and total_weight_container > 0:
                    cog_per_unit = (sum_expenses / Decimal('2') / total_qty_container) + \
                                  (sum_expenses / Decimal('2') / total_weight_container * item_weight)
                elif total_qty_container > 0:
                    cog_per_unit = sum_expenses / total_qty_container
                else:
                    cog_per_unit = Decimal('0')
                
                # Item cost per unit = unit_price + COG per unit
                item_cost_per_unit = item_pi.unit_price + cog_per_unit
                total_cost = item_cost_per_unit * item_pi.quantity
                
                containers_data.append({
                    'container_number': container.container_number,
                    'container_date': container.date.isoformat() if container.date else None,
                    'container_currency': container.currency,
                    'quantity': float(item_pi.quantity),
                    'unit_price': float(item_pi.unit_price),
                    'expense1_original': float(expense1_original),
                    'expense1_currency': expense1_currency,
                    'expense1_in_container_currency': float(expense1_in_container_currency),
                    'expense2_original': float(expense2_original),
                    'expense2_currency': expense2_currency,
                    'expense2_in_container_currency': float(expense2_in_container_currency),
                    'expense3_original': float(expense3_original),
                    'expense3_currency': expense3_currency,
                    'expense3_in_container_currency': float(expense3_in_container_currency),
                    'total_expenses_in_container_currency': float(sum_expenses),
                    'cog_per_unit': float(cog_per_unit),
                    'item_cost_per_unit': float(item_cost_per_unit),
                    'total_cost': float(total_cost)
                })
                
                total_cost_all_containers += total_cost
                total_quantity_all_containers += item_pi.quantity
            
            # Calculate average cost per unit
            average_cost_per_unit = Decimal('0')
            if total_quantity_all_containers > 0:
                average_cost_per_unit = total_cost_all_containers / total_quantity_all_containers
            
            # Calculate stock value based on method
            if calculation_method == 'FIFO':
                # For FIFO, calculate from batches
                batches = InventoryBatch.query.filter_by(
                    market_id=market_id,
                    item_id=item.id
                ).filter(
                    InventoryBatch.available_quantity > 0
                ).all()
                
                stock_value = sum(batch.available_quantity * batch.cost_per_unit for batch in batches)
                
                # Apply adjustments using last batch cost
                if adjustment_qty != 0:
                    last_batch = InventoryBatch.query.filter_by(
                        market_id=market_id,
                        item_id=item.id
                    ).order_by(
                        InventoryBatch.purchase_date.desc(),
                        InventoryBatch.id.desc()
                    ).first()
                    if last_batch:
                        stock_value += adjustment_qty * last_batch.cost_per_unit
                
                # Update containers_data with batch info for FIFO
                containers_data = []
                for batch in batches:
                    container = batch.container
                    if not container:
                        continue
                    
                    # Get expense info from container
                    expense1_in_container_currency = Decimal('0')
                    expense1_original = Decimal('0')
                    expense1_currency = container.currency
                    if container.expense1_amount and container.expense1_amount > 0:
                        expense1_original = container.expense1_amount
                        expense1_currency = container.expense1_currency
                        if container.expense1_currency == container.currency:
                            expense1_in_container_currency = container.expense1_amount
                        else:
                            expense1_base = container.expense1_amount * (container.expense1_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense1_in_container_currency = expense1_base / container_rate
                    
                    expense2_in_container_currency = Decimal('0')
                    expense2_original = Decimal('0')
                    expense2_currency = container.currency
                    if container.expense2_amount and container.expense2_amount > 0:
                        expense2_original = container.expense2_amount
                        expense2_currency = container.expense2_currency
                        if container.expense2_currency == container.currency:
                            expense2_in_container_currency = container.expense2_amount
                        else:
                            expense2_base = container.expense2_amount * (container.expense2_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense2_in_container_currency = expense2_base / container_rate
                    
                    expense3_in_container_currency = Decimal('0')
                    expense3_original = Decimal('0')
                    expense3_currency = container.currency
                    if container.expense3_amount and container.expense3_amount > 0:
                        expense3_original = container.expense3_amount
                        expense3_currency = container.expense3_currency
                        if container.expense3_currency == container.currency:
                            expense3_in_container_currency = container.expense3_amount
                        else:
                            expense3_base = container.expense3_amount * (container.expense3_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense3_in_container_currency = expense3_base / container_rate
                    
                    sum_expenses = expense1_in_container_currency + expense2_in_container_currency + expense3_in_container_currency
                    
                    containers_data.append({
                        'container_number': container.container_number,
                        'container_date': container.date.isoformat() if container.date else None,
                        'container_currency': batch.currency,
                        'quantity': float(batch.available_quantity),
                        'unit_price': float(batch.unit_price),
                        'expense1_original': float(expense1_original),
                        'expense1_currency': expense1_currency,
                        'expense1_in_container_currency': float(expense1_in_container_currency),
                        'expense2_original': float(expense2_original),
                        'expense2_currency': expense2_currency,
                        'expense2_in_container_currency': float(expense2_in_container_currency),
                        'expense3_original': float(expense3_original),
                        'expense3_currency': expense3_currency,
                        'expense3_in_container_currency': float(expense3_in_container_currency),
                        'total_expenses_in_container_currency': float(sum_expenses),
                        'cog_per_unit': float(batch.cog_per_unit),
                        'item_cost_per_unit': float(batch.cost_per_unit),
                        'total_cost': float(batch.available_quantity * batch.cost_per_unit)
                    })
                
                # Recalculate totals from batches
                total_cost_all_containers = sum(batch.available_quantity * batch.cost_per_unit for batch in batches)
                total_quantity_all_containers = sum(batch.available_quantity for batch in batches)
                if total_quantity_all_containers > 0:
                    average_cost_per_unit = total_cost_all_containers / total_quantity_all_containers
            else:
                # Average Cost method
                stock_value = available_qty * average_cost_per_unit
            
            supplier_items.append({
                'item_code': item.code,
                'item_name': item.name,
                'item_weight': float(item.weight) if item.weight else 0.0,
                'currency': supplier.currency,
                'purchased_quantity': float(purchased_qty),
                'sold_quantity': float(sold_qty),
                'available_quantity': float(available_qty),
                'containers': containers_data,
                'total_cost_all_containers': float(total_cost_all_containers),
                'total_quantity_all_containers': float(total_quantity_all_containers),
                'average_cost_per_unit': float(average_cost_per_unit),
                'stock_value': float(stock_value)
            })
        
        if supplier_items:
            suppliers_data.append({
                'supplier_name': supplier.name,
                'supplier_currency': supplier.currency,
                'items': supplier_items
            })
    
    return jsonify({
        'success': True,
        'data': suppliers_data
    })

@bp.route('/stock-value-details/export', methods=['GET'])
@login_required
def export_stock_value_details():
    """Export stock value details report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    item_id = request.args.get('item_id', type=int)
    
    # Get market calculation method
    market = Market.query.get(market_id)
    calculation_method = getattr(market, 'calculation_method', 'Average') if market else 'Average'
    
    # Get all suppliers
    suppliers = Company.query.filter_by(market_id=market_id, category='Supplier').all()
    
    # Build export data - flatten structure for Excel
    export_data = []
    
    for supplier in suppliers:
        # Get items for this supplier
        items_query = Item.query.filter_by(market_id=market_id, supplier_id=supplier.id)
        if item_id:
            items_query = items_query.filter_by(id=item_id)
        items = items_query.all()
        
        if not items:
            continue
        
        for item in items:
            # Get purchased quantity
            purchased_qty = db.session.query(func.coalesce(func.sum(PurchaseItem.quantity), 0)).join(
                PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
            ).filter(
                PurchaseContainer.market_id == market_id,
                PurchaseItem.item_id == item.id
            ).scalar() or Decimal('0')
            
            # Get sold quantity
            sold_qty = db.session.query(func.coalesce(func.sum(SaleItem.quantity), 0)).join(
                Sale, SaleItem.sale_id == Sale.id
            ).filter(
                Sale.market_id == market_id,
                SaleItem.item_id == item.id
            ).scalar() or Decimal('0')
            
            # Get inventory adjustments
            adjustment_qty = db.session.query(func.sum(
                case(
                    (InventoryAdjustment.adjustment_type == 'Increase', InventoryAdjustment.quantity),
                    else_=-InventoryAdjustment.quantity
                )
            )).filter(
                InventoryAdjustment.market_id == market_id,
                InventoryAdjustment.item_id == item.id
            ).scalar() or Decimal('0')
            
            available_qty = purchased_qty - sold_qty + adjustment_qty
            
            # Get containers for this item
            purchase_items = PurchaseItem.query.join(
                PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
            ).filter(
                PurchaseContainer.market_id == market_id,
                PurchaseItem.item_id == item.id
            ).all()
            
            # Group by container
            containers_map = {}
            for pi in purchase_items:
                container_id = pi.container_id
                if container_id not in containers_map:
                    containers_map[container_id] = []
                containers_map[container_id].append(pi)
            
            # Calculate stock value
            if calculation_method == 'FIFO':
                batches = InventoryBatch.query.filter_by(
                    market_id=market_id,
                    item_id=item.id
                ).filter(
                    InventoryBatch.available_quantity > 0
                ).all()
                
                stock_value = sum(batch.available_quantity * batch.cost_per_unit for batch in batches)
                
                if adjustment_qty != 0:
                    last_batch = InventoryBatch.query.filter_by(
                        market_id=market_id,
                        item_id=item.id
                    ).order_by(
                        InventoryBatch.purchase_date.desc(),
                        InventoryBatch.id.desc()
                    ).first()
                    if last_batch:
                        stock_value += adjustment_qty * last_batch.cost_per_unit
                
                # Calculate average cost from batches
                total_cost = sum(batch.available_quantity * batch.cost_per_unit for batch in batches)
                total_qty = sum(batch.available_quantity for batch in batches)
                avg_cost = total_cost / total_qty if total_qty > 0 else Decimal('0')
            else:
                # Average Cost - calculate from purchase items
                total_cost = Decimal('0')
                total_qty = Decimal('0')
                
                for container_id, purchase_items_list in containers_map.items():
                    container = PurchaseContainer.query.get(container_id)
                    if not container:
                        continue
                    
                    # Calculate expenses
                    expense1_in_container_currency = Decimal('0')
                    if container.expense1_amount and container.expense1_amount > 0:
                        if container.expense1_currency == container.currency:
                            expense1_in_container_currency = container.expense1_amount
                        else:
                            expense1_base = container.expense1_amount * (container.expense1_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense1_in_container_currency = expense1_base / container_rate
                    
                    expense2_in_container_currency = Decimal('0')
                    if container.expense2_amount and container.expense2_amount > 0:
                        if container.expense2_currency == container.currency:
                            expense2_in_container_currency = container.expense2_amount
                        else:
                            expense2_base = container.expense2_amount * (container.expense2_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense2_in_container_currency = expense2_base / container_rate
                    
                    expense3_in_container_currency = Decimal('0')
                    if container.expense3_amount and container.expense3_amount > 0:
                        if container.expense3_currency == container.currency:
                            expense3_in_container_currency = container.expense3_amount
                        else:
                            expense3_base = container.expense3_amount * (container.expense3_exchange_rate or 1)
                            container_rate = container.exchange_rate or 1
                            if container_rate > 0:
                                expense3_in_container_currency = expense3_base / container_rate
                    
                    sum_expenses = expense1_in_container_currency + expense2_in_container_currency + expense3_in_container_currency
                    
                    total_qty_container = sum(pi.quantity for pi in purchase_items_list)
                    total_weight_container = sum((pi.item.weight or Decimal('0')) * pi.quantity for pi in purchase_items_list)
                    
                    item_pi = next((pi for pi in purchase_items_list if pi.item_id == item.id), None)
                    if item_pi:
                        item_weight = item.weight or Decimal('0')
                        
                        if total_qty_container > 0 and total_weight_container > 0:
                            cog_per_unit = (sum_expenses / Decimal('2') / total_qty_container) + \
                                          (sum_expenses / Decimal('2') / total_weight_container * item_weight)
                        elif total_qty_container > 0:
                            cog_per_unit = sum_expenses / total_qty_container
                        else:
                            cog_per_unit = Decimal('0')
                        
                        item_cost_per_unit = item_pi.unit_price + cog_per_unit
                        total_cost += item_cost_per_unit * item_pi.quantity
                        total_qty += item_pi.quantity
                
                avg_cost = total_cost / total_qty if total_qty > 0 else Decimal('0')
                stock_value = available_qty * avg_cost
            
            # Add row for each item
            export_data.append({
                'Supplier': supplier.name,
                'Supplier Currency': supplier.currency,
                'Item Code': item.code,
                'Item Name': item.name,
                'Item Weight': float(item.weight) if item.weight else 0.0,
                'Purchased Quantity': float(purchased_qty),
                'Sold Quantity': float(sold_qty),
                'Available Quantity': float(available_qty),
                'Average Cost Per Unit': float(avg_cost),
                'Stock Value': float(stock_value),
                'Calculation Method': calculation_method
            })
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Stock Value Details')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Stock Value Details']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
    
    output.seek(0)
    filename = f'stock_value_details_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@bp.route('/item-statement', methods=['GET'])
@login_required
def get_item_statement():
    """Get item statement report"""
    return jsonify({'error': 'Endpoint not yet implemented'}), 501

@bp.route('/virtual-purchase-profit', methods=['POST'])
@login_required
def virtual_purchase_profit():
    """Virtual purchase profit report"""
    return jsonify({'error': 'Endpoint not yet implemented'}), 501

@bp.route('/average-sale-price', methods=['GET'])
@login_required
def get_average_sale_price():
    """Get average sale price report - Average Sale Price = Total Revenue ÷ Total Quantity per item"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    supplier_id = request.args.get('supplier_id', type=int)
    customer_id = request.args.get('customer_id', type=int)
    item_id = request.args.get('item_id', type=int)

    query = db.session.query(SaleItem, Sale, Item).join(
        Sale, SaleItem.sale_id == Sale.id
    ).join(
        Item, SaleItem.item_id == Item.id
    ).filter(Sale.market_id == market_id)

    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if supplier_id:
        query = query.filter(Item.supplier_id == supplier_id)
    if customer_id:
        query = query.filter(Sale.customer_id == customer_id)
    if item_id:
        query = query.filter(SaleItem.item_id == item_id)

    sale_items = query.order_by(Sale.date.asc(), Sale.id.asc()).all()

    # Aggregate by item
    items_data = {}
    for si, sale, item in sale_items:
        iid = item.id
        if iid not in items_data:
            supplier = item.supplier
            items_data[iid] = {
                'item_id': iid,
                'item_code': item.code,
                'item_name': item.name,
                'supplier_name': supplier.name if supplier else None,
                'total_quantity_sold': Decimal('0'),
                'total_revenue': Decimal('0'),
                'sales': []
            }

        items_data[iid]['total_quantity_sold'] += si.quantity
        items_data[iid]['total_revenue'] += si.total_price
        items_data[iid]['sales'].append({
            'date': sale.date.isoformat(),
            'invoice_number': sale.invoice_number,
            'customer_name': sale.customer.name if sale.customer else 'Unknown',
            'customer_currency': sale.customer.currency if sale.customer else 'CFA',
            'quantity': float(si.quantity),
            'unit_price': float(si.unit_price),
            'total_price': float(si.total_price)
        })

    items_list = []
    for item_data in items_data.values():
        total_qty = float(item_data['total_quantity_sold'])
        total_rev = float(item_data['total_revenue'])
        avg_price = total_rev / total_qty if total_qty > 0 else 0

        items_list.append({
            'item_id': item_data['item_id'],
            'item_code': item_data['item_code'],
            'item_name': item_data['item_name'],
            'supplier_name': item_data['supplier_name'],
            'average_sale_price': round(avg_price, 2),
            'total_quantity_sold': total_qty,
            'total_revenue': round(total_rev, 2),
            'transaction_count': len(item_data['sales']),
            'sales': item_data['sales']
        })

    # Sort by item_code
    items_list.sort(key=lambda x: (x['item_code'] or '', x['item_name'] or ''))

    return jsonify({
        'items': items_list,
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'supplier_id': supplier_id,
            'customer_id': customer_id,
            'item_id': item_id
        }
    })

def _get_average_sale_price_data(market_id, start_date, end_date, supplier_id, customer_id, item_id):
    """Shared logic for average sale price report and export"""
    query = db.session.query(SaleItem, Sale, Item).join(
        Sale, SaleItem.sale_id == Sale.id
    ).join(
        Item, SaleItem.item_id == Item.id
    ).filter(Sale.market_id == market_id)

    if start_date:
        query = query.filter(Sale.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Sale.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if supplier_id:
        query = query.filter(Item.supplier_id == supplier_id)
    if customer_id:
        query = query.filter(Sale.customer_id == customer_id)
    if item_id:
        query = query.filter(SaleItem.item_id == item_id)

    sale_items = query.order_by(Sale.date.asc(), Sale.id.asc()).all()

    items_data = {}
    for si, sale, item in sale_items:
        iid = item.id
        if iid not in items_data:
            supplier = item.supplier
            items_data[iid] = {
                'item_id': iid,
                'item_code': item.code,
                'item_name': item.name,
                'supplier_name': supplier.name if supplier else None,
                'total_quantity_sold': Decimal('0'),
                'total_revenue': Decimal('0'),
            }

        items_data[iid]['total_quantity_sold'] += si.quantity
        items_data[iid]['total_revenue'] += si.total_price

    items_list = []
    for item_data in items_data.values():
        total_qty = float(item_data['total_quantity_sold'])
        total_rev = float(item_data['total_revenue'])
        avg_price = total_rev / total_qty if total_qty > 0 else 0
        items_list.append({
            'item_code': item_data['item_code'],
            'item_name': item_data['item_name'],
            'supplier_name': item_data['supplier_name'] or '',
            'average_sale_price': round(avg_price, 2),
            'total_quantity_sold': total_qty,
            'total_revenue': round(total_rev, 2),
        })

    items_list.sort(key=lambda x: (x['item_code'] or '', x['item_name'] or ''))
    return items_list

@bp.route('/average-sale-price/export', methods=['GET'])
@login_required
def export_average_sale_price():
    """Export Average Sale Price report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    supplier_id = request.args.get('supplier_id', type=int)
    customer_id = request.args.get('customer_id', type=int)
    item_id = request.args.get('item_id', type=int)

    items_list = _get_average_sale_price_data(market_id, start_date, end_date, supplier_id, customer_id, item_id)

    export_data = []
    for row in items_list:
        export_data.append({
            'Item Code': row['item_code'],
            'Item Name': row['item_name'],
            'Supplier': row['supplier_name'],
            'Average Sale Price': row['average_sale_price'],
            'Total Quantity Sold': row['total_quantity_sold'],
            'Total Revenue': row['total_revenue'],
        })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Average Sale Price')

        worksheet = writer.sheets['Average Sale Price']
        for idx, col in enumerate(df.columns):
            max_length = len(str(col)) if df.empty else max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)

    output.seek(0)
    filename = f'average_sale_price_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@bp.route('/last-purchase-price', methods=['GET'])
@login_required
def get_last_purchase_price():
    """Last purchase price for all items - the unit price from the most recent purchase per item"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    supplier_id = request.args.get('supplier_id', type=int)
    item_id = request.args.get('item_id', type=int)

    # Get items that have purchases (with optional filters)
    items_query = Item.query.filter_by(market_id=market_id)
    if supplier_id:
        items_query = items_query.filter(Item.supplier_id == supplier_id)
    if item_id:
        items_query = items_query.filter(Item.id == item_id)

    items_list = []
    for item in items_query.all():
        # Get most recent purchase for this item
        last_purchase = db.session.query(PurchaseItem, PurchaseContainer).join(
            PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
        ).filter(
            PurchaseContainer.market_id == market_id,
            PurchaseItem.item_id == item.id
        ).order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).first()

        if not last_purchase:
            continue

        pi, container = last_purchase
        supplier = container.supplier if container else None

        items_list.append({
            'item_id': item.id,
            'item_code': item.code,
            'item_name': item.name,
            'supplier_name': supplier.name if supplier else None,
            'last_purchase_price': float(pi.unit_price),
            'last_purchase_date': container.date.isoformat() if container.date else None,
            'container_number': container.container_number if container else None,
            'quantity': float(pi.quantity),
            'total_price': float(pi.total_price),
            'currency': container.currency if container else None
        })

    items_list.sort(key=lambda x: (x['item_code'] or '', x['item_name'] or ''))

    return jsonify({
        'items': items_list,
        'filters': {'supplier_id': supplier_id, 'item_id': item_id}
    })

@bp.route('/last-purchase-price/export', methods=['GET'])
@login_required
def export_last_purchase_price():
    """Export Last Purchase Price report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    supplier_id = request.args.get('supplier_id', type=int)
    item_id = request.args.get('item_id', type=int)

    items_query = Item.query.filter_by(market_id=market_id)
    if supplier_id:
        items_query = items_query.filter(Item.supplier_id == supplier_id)
    if item_id:
        items_query = items_query.filter(Item.id == item_id)

    export_data = []
    for item in items_query.all():
        last_purchase = db.session.query(PurchaseItem, PurchaseContainer).join(
            PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
        ).filter(
            PurchaseContainer.market_id == market_id,
            PurchaseItem.item_id == item.id
        ).order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).first()

        if not last_purchase:
            continue

        pi, container = last_purchase
        supplier = container.supplier if container else None

        export_data.append({
            'Item Code': item.code,
            'Item Name': item.name,
            'Supplier': supplier.name if supplier else '',
            'Last Purchase Price': float(pi.unit_price),
            'Last Purchase Date': container.date.isoformat() if container.date else '',
            'Container Number': container.container_number if container else '',
            'Quantity': float(pi.quantity),
            'Total Price': float(pi.total_price),
            'Currency': container.currency if container else '',
        })

    export_data.sort(key=lambda x: (x['Item Code'] or '', x['Item Name'] or ''))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Last Purchase Price')

        worksheet = writer.sheets['Last Purchase Price']
        for idx, col in enumerate(df.columns):
            max_length = len(str(col)) if df.empty else max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)

    output.seek(0)
    filename = f'last_purchase_price_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@bp.route('/average-last-n-sales', methods=['GET'])
@login_required
def get_average_last_n_sales():
    """Average of last N sales per item (N=10, or 9, 8, ... if fewer available).
    For each item: take the last 10 sale line items (most recent), or all if fewer than 10.
    Average = Total Revenue ÷ Total Quantity from those transactions."""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    supplier_id = request.args.get('supplier_id', type=int)
    item_id = request.args.get('item_id', type=int)
    max_n = 10

    # Get all items that have sales (optional: filter by supplier/item)
    items_query = Item.query.filter_by(market_id=market_id)
    if supplier_id:
        items_query = items_query.filter(Item.supplier_id == supplier_id)
    if item_id:
        items_query = items_query.filter(Item.id == item_id)

    items_list = []
    for item in items_query.all():
        # Get sale items for this item, most recent first
        q = db.session.query(SaleItem, Sale).join(
            Sale, SaleItem.sale_id == Sale.id
        ).filter(
            Sale.market_id == market_id,
            SaleItem.item_id == item.id
        ).options(joinedload(Sale.customer)).order_by(Sale.date.desc(), Sale.id.desc())

        rows = q.limit(max_n).all()

        if not rows:
            continue

        total_qty = sum(float(si.quantity) for si, _ in rows)
        total_rev = sum(float(si.total_price) for si, _ in rows)
        n_used = len(rows)
        avg_price = total_rev / total_qty if total_qty > 0 else 0

        sales_detail = []
        for si, sale in rows:
            sales_detail.append({
                'date': sale.date.isoformat(),
                'invoice_number': sale.invoice_number,
                'customer_name': sale.customer.name if sale.customer else 'Unknown',
                'customer_currency': sale.customer.currency if sale.customer else 'CFA',
                'quantity': float(si.quantity),
                'unit_price': float(si.unit_price),
                'total_price': float(si.total_price)
            })

        supplier = item.supplier
        items_list.append({
            'item_id': item.id,
            'item_code': item.code,
            'item_name': item.name,
            'supplier_name': supplier.name if supplier else None,
            'average_sale_price': round(avg_price, 2),
            'total_quantity_sold': total_qty,
            'total_revenue': round(total_rev, 2),
            'sales_used': n_used,
            'sales': sales_detail
        })

    items_list.sort(key=lambda x: (x['item_code'] or '', x['item_name'] or ''))

    return jsonify({
        'items': items_list,
        'max_n': max_n,
        'filters': {'supplier_id': supplier_id, 'item_id': item_id}
    })

@bp.route('/average-last-n-sales/export', methods=['GET'])
@login_required
def export_average_last_n_sales():
    """Export Average of Last N Sales report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    supplier_id = request.args.get('supplier_id', type=int)
    item_id = request.args.get('item_id', type=int)
    max_n = 10

    items_query = Item.query.filter_by(market_id=market_id)
    if supplier_id:
        items_query = items_query.filter(Item.supplier_id == supplier_id)
    if item_id:
        items_query = items_query.filter(Item.id == item_id)

    export_data = []
    for item in items_query.all():
        q = db.session.query(SaleItem, Sale).join(
            Sale, SaleItem.sale_id == Sale.id
        ).filter(
            Sale.market_id == market_id,
            SaleItem.item_id == item.id
        ).order_by(Sale.date.desc(), Sale.id.desc())

        rows = q.limit(max_n).all()
        if not rows:
            continue

        total_qty = sum(float(si.quantity) for si, _ in rows)
        total_rev = sum(float(si.total_price) for si, _ in rows)
        n_used = len(rows)
        avg_price = total_rev / total_qty if total_qty > 0 else 0

        supplier = item.supplier
        export_data.append({
            'Item Code': item.code,
            'Item Name': item.name,
            'Supplier': supplier.name if supplier else '',
            'Sales Used (N)': n_used,
            'Average Sale Price': round(avg_price, 2),
            'Total Quantity': total_qty,
            'Total Revenue': round(total_rev, 2),
        })

    export_data.sort(key=lambda x: (x['Item Code'] or '', x['Item Name'] or ''))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Avg Last N Sales')

        worksheet = writer.sheets['Avg Last N Sales']
        for idx, col in enumerate(df.columns):
            max_length = len(str(col)) if df.empty else max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)

    output.seek(0)
    filename = f'avg_last_n_sales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@bp.route('/safe-out', methods=['GET'])
@login_required
def get_safe_out_report():
    """Get Safe Out Report - combines payments (Out) and general expenses"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get all Out payments (eager load company and sale to avoid N+1 queries)
    payments_query = Payment.query.options(
        joinedload(Payment.company),
        joinedload(Payment.sale)
    ).filter_by(
        market_id=market_id,
        payment_type='Out'
    )
    
    if start_date:
        payments_query = payments_query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        payments_query = payments_query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    payments = payments_query.order_by(Payment.date.asc(), Payment.id.asc()).all()
    
    # Get all general expenses
    expenses_query = GeneralExpense.query.filter_by(market_id=market_id)
    
    if start_date:
        expenses_query = expenses_query.filter(GeneralExpense.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        expenses_query = expenses_query.filter(GeneralExpense.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    expenses = expenses_query.order_by(GeneralExpense.date.asc(), GeneralExpense.id.asc()).all()
    
    # Combine into a single list with type indicator
    transactions = []
    
    for payment in payments:
        amount_base = payment.amount_base_currency_stored if payment.amount_base_currency_stored else payment.amount * payment.exchange_rate
        transactions.append({
            'id': payment.id,
            'date': payment.date.isoformat(),
            'type': 'Payment',
            'description': payment.company.name if payment.company else 'Unknown',
            'category': 'Loan' if payment.loan else payment.company.category if payment.company else 'Unknown',
            'invoice_number': payment.sale.invoice_number if payment.sale else None,
            'amount': float(payment.amount),
            'currency': payment.currency,
            'exchange_rate': float(payment.exchange_rate),
            'amount_base_currency': float(amount_base),
            'notes': payment.notes,
            'loan': payment.loan
        })
    
    for expense in expenses:
        amount_base = expense.amount_base_currency
        transactions.append({
            'id': expense.id,
            'date': expense.date.isoformat(),
            'type': 'Expense',
            'description': expense.description,
            'category': expense.category,
            'invoice_number': None,
            'amount': float(expense.amount),
            'currency': expense.currency,
            'exchange_rate': float(expense.exchange_rate),
            'amount_base_currency': float(amount_base),
            'notes': None,
            'loan': False
        })
    
    # Sort by date
    transactions.sort(key=lambda x: (x['date'], x['id']))
    
    # Calculate totals
    total_payments = sum(t['amount_base_currency'] for t in transactions if t['type'] == 'Payment')
    total_expenses = sum(t['amount_base_currency'] for t in transactions if t['type'] == 'Expense')
    total_out = total_payments + total_expenses
    
    return jsonify({
        'transactions': transactions,
        'totals': {
            'total_payments': float(total_payments),
            'total_expenses': float(total_expenses),
            'total_out': float(total_out),
            'count': len(transactions)
        }
    })

@bp.route('/safe-out/export', methods=['GET'])
@login_required
def export_safe_out_report():
    """Export Safe Out Report to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Get data using the same logic as the report endpoint (eager load to avoid N+1)
    payments_query = Payment.query.options(
        joinedload(Payment.company),
        joinedload(Payment.sale)
    ).filter_by(
        market_id=market_id,
        payment_type='Out'
    )
    
    if start_date:
        payments_query = payments_query.filter(Payment.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        payments_query = payments_query.filter(Payment.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    payments = payments_query.order_by(Payment.date.asc(), Payment.id.asc()).all()
    
    expenses_query = GeneralExpense.query.filter_by(market_id=market_id)
    
    if start_date:
        expenses_query = expenses_query.filter(GeneralExpense.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        expenses_query = expenses_query.filter(GeneralExpense.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    expenses = expenses_query.order_by(GeneralExpense.date.asc(), GeneralExpense.id.asc()).all()
    
    # Build export data
    export_data = []
    
    for payment in payments:
        amount_base = payment.amount_base_currency_stored if payment.amount_base_currency_stored else payment.amount * payment.exchange_rate
        export_data.append({
            'Date': payment.date.isoformat(),
            'Type': 'Payment',
            'Description/Company': payment.company.name if payment.company else 'Unknown',
            'Category': 'Loan' if payment.loan else payment.company.category if payment.company else 'Unknown',
            'Invoice Number': payment.sale.invoice_number if payment.sale else '',
            'Amount': float(payment.amount),
            'Currency': payment.currency,
            'Exchange Rate': float(payment.exchange_rate),
            'Amount (Base Currency)': float(amount_base),
            'Notes': payment.notes or ''
        })
    
    for expense in expenses:
        amount_base = expense.amount_base_currency
        export_data.append({
            'Date': expense.date.isoformat(),
            'Type': 'Expense',
            'Description/Company': expense.description,
            'Category': expense.category,
            'Invoice Number': '',
            'Amount': float(expense.amount),
            'Currency': expense.currency,
            'Exchange Rate': float(expense.exchange_rate),
            'Amount (Base Currency)': float(amount_base),
            'Notes': ''
        })
    
    # Sort by date
    export_data.sort(key=lambda x: (x['Date'], x.get('ID', 0)))
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df = pd.DataFrame(export_data)
        df.to_excel(writer, index=False, sheet_name='Safe Out Report')
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Safe Out Report']
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
    
    output.seek(0)
    filename = f'safe_out_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

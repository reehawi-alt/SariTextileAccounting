"""
Purchases API endpoints
"""
from flask import Blueprint, request, jsonify, session, send_file
from flask_login import login_required
from models import db, PurchaseContainer, PurchaseItem, Item, Market, Company, SafeTransaction
from decimal import Decimal
from datetime import datetime
import pandas as pd
from io import BytesIO

bp = Blueprint('purchases', __name__)

@bp.route('/containers', methods=['GET'])
@login_required
def get_containers():
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    supplier_id = request.args.get('supplier_id', type=int)
    
    query = PurchaseContainer.query.filter_by(market_id=market_id)
    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)
    
    containers = query.order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).all()
    
    return jsonify([{
        'id': c.id,
        'container_number': c.container_number,
        'supplier_id': c.supplier_id,
        'supplier_name': c.supplier.name,
        'currency': c.currency,
        'exchange_rate': float(c.exchange_rate),
        'date': c.date.isoformat(),
        'total_amount': float(c.total_amount),
        'notes': c.notes
    } for c in containers])

@bp.route('/containers', methods=['POST'])
@login_required
def create_container():
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    data = request.json
    
    # Process expense values
    expense1_amount_val = data.get('expense1_amount', 0) or 0
    expense1_amount = Decimal(str(expense1_amount_val)) if expense1_amount_val else None
    # Expense1 always uses container currency and exchange rate
    expense1_currency = data.get('currency')  # Use container currency
    expense1_exchange_rate = Decimal(str(data['exchange_rate']))  # Use container exchange rate
    
    expense2_amount_val = data.get('expense2_amount', 0) or 0
    expense2_amount = Decimal(str(expense2_amount_val)) if expense2_amount_val else None
    expense2_service_company_id = data.get('expense2_service_company_id')
    expense2_currency = data.get('expense2_currency') or None
    expense2_exchange_rate_val = data.get('expense2_exchange_rate')
    expense2_exchange_rate = Decimal(str(expense2_exchange_rate_val)) if expense2_exchange_rate_val is not None else None
    
    expense3_amount_val = data.get('expense3_amount', 0) or 0
    expense3_amount = Decimal(str(expense3_amount_val)) if expense3_amount_val else None
    expense3_currency = data.get('expense3_currency') or None
    expense3_exchange_rate_val = data.get('expense3_exchange_rate')
    expense3_exchange_rate = Decimal(str(expense3_exchange_rate_val)) if expense3_exchange_rate_val is not None else None
    
    container = PurchaseContainer(
        market_id=market_id,
        container_number=data['container_number'],
        supplier_id=data['supplier_id'],
        currency=data['currency'],
        exchange_rate=Decimal(str(data['exchange_rate'])),
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        notes=data.get('notes', ''),
        # Expenses
        expense1_amount=expense1_amount,
        expense1_currency=expense1_currency,
        expense1_exchange_rate=expense1_exchange_rate,
        expense2_amount=expense2_amount,
        expense2_service_company_id=expense2_service_company_id,
        expense2_currency=expense2_currency,
        expense2_exchange_rate=expense2_exchange_rate,
        expense3_amount=expense3_amount,
        expense3_currency=expense3_currency,
        expense3_exchange_rate=expense3_exchange_rate
    )
    
    db.session.add(container)
    db.session.flush()
    
    # Add items
    for item_data in data.get('items', []):
        # Ensure item belongs to this supplier
        item_obj = Item.query.get(item_data['item_id'])
        if item_obj:
            # Update item's supplier_id if not set or different
            if not item_obj.supplier_id or item_obj.supplier_id != container.supplier_id:
                item_obj.supplier_id = container.supplier_id
        
        item = PurchaseItem(
            container_id=container.id,
            item_id=item_data['item_id'],
            quantity=Decimal(str(item_data['quantity'])),
            unit_price=Decimal(str(item_data['unit_price'])),
            total_price=Decimal(str(item_data['quantity'])) * Decimal(str(item_data['unit_price']))
        )
        db.session.add(item)
    
    # Record expense3 (cash expense) in safe if exists
    if container.expense3_amount and container.expense3_amount > 0:
        market = Market.query.get(market_id)
        last_transaction = SafeTransaction.query.filter_by(market_id=market_id).order_by(
            SafeTransaction.date.desc(), SafeTransaction.id.desc()
        ).first()
        
        balance_before = last_transaction.balance_after if last_transaction else Decimal('0')
        balance_after = balance_before - container.expense3_base_currency
        
        # Calculate exact base currency amount
        expense3_rate = container.expense3_exchange_rate or container.exchange_rate
        exact_base_amount = container.expense3_amount * expense3_rate
        
        safe_transaction = SafeTransaction(
            market_id=market_id,
            transaction_type='Outflow',
            amount=container.expense3_amount,
            currency=container.expense3_currency or container.currency,
            exchange_rate=expense3_rate,
            amount_base_currency_stored=exact_base_amount,  # Store exact value directly
            date=container.date,
            description=f'Container {container.container_number} - Expense 3 (Cash Expense)',
            balance_after=balance_after
        )
        db.session.add(safe_transaction)
    
    db.session.commit()
    
    # Create inventory batches if FIFO is active
    from models import Market
    market = Market.query.get(market_id)
    if market and getattr(market, 'calculation_method', 'Average') == 'FIFO':
        from api.fifo_calculations import create_inventory_batches_for_container
        create_inventory_batches_for_container(container.id)
    
    return jsonify({
        'id': container.id,
        'container_number': container.container_number,
        'supplier_id': container.supplier_id,
        'supplier_name': container.supplier.name,
        'currency': container.currency,
        'exchange_rate': float(container.exchange_rate),
        'date': container.date.isoformat(),
        'total_amount': float(container.total_amount),
        'notes': container.notes,
        'expense1_amount': float(container.expense1_amount or 0),
        'expense1_currency': container.expense1_currency,
        'expense1_exchange_rate': float(container.expense1_exchange_rate) if container.expense1_exchange_rate is not None else None,
        'expense2_amount': float(container.expense2_amount or 0),
        'expense2_service_company_id': container.expense2_service_company_id,
        'expense2_service_company_name': container.expense_service_company.name if container.expense_service_company else None,
        'expense2_currency': container.expense2_currency,
        'expense2_exchange_rate': float(container.expense2_exchange_rate) if container.expense2_exchange_rate is not None else None,
        'expense3_amount': float(container.expense3_amount or 0),
        'expense3_currency': container.expense3_currency,
        'expense3_exchange_rate': float(container.expense3_exchange_rate) if container.expense3_exchange_rate is not None else None
    }), 201

@bp.route('/containers/<int:container_id>', methods=['GET'])
@login_required
def get_container(container_id):
    market_id = session.get('current_market_id')
    container = PurchaseContainer.query.filter_by(id=container_id, market_id=market_id).first()
    
    if not container:
        return jsonify({'error': 'Container not found'}), 404
    
    items = [{
        'id': i.id,
        'item_id': i.item_id,
        'item_code': i.item.code,
        'item_name': i.item.name,
        'quantity': float(i.quantity),
        'unit_price': float(i.unit_price),
        'total_price': float(i.total_price)
    } for i in container.items]
    
    return jsonify({
        'id': container.id,
        'container_number': container.container_number,
        'supplier_id': container.supplier_id,
        'supplier_name': container.supplier.name,
        'currency': container.currency,
        'exchange_rate': float(container.exchange_rate),
        'date': container.date.isoformat(),
        'total_amount': float(container.total_amount),
        'notes': container.notes,
        'expense1_amount': float(container.expense1_amount or 0),
        'expense1_currency': container.expense1_currency,
        'expense1_exchange_rate': float(container.expense1_exchange_rate) if container.expense1_exchange_rate is not None else None,
        'expense2_amount': float(container.expense2_amount or 0),
        'expense2_service_company_id': container.expense2_service_company_id,
        'expense2_service_company_name': container.expense_service_company.name if container.expense_service_company else None,
        'expense2_currency': container.expense2_currency,
        'expense2_exchange_rate': float(container.expense2_exchange_rate) if container.expense2_exchange_rate is not None else None,
        'expense3_amount': float(container.expense3_amount or 0),
        'expense3_currency': container.expense3_currency,
        'expense3_exchange_rate': float(container.expense3_exchange_rate) if container.expense3_exchange_rate is not None else None,
        'items': items
    })

@bp.route('/containers/<int:container_id>', methods=['PUT'])
@login_required
def update_container(container_id):
    try:
        market_id = session.get('current_market_id')
        container = PurchaseContainer.query.filter_by(id=container_id, market_id=market_id).first()
        
        if not container:
            return jsonify({'error': 'Container not found'}), 404
        
        data = request.json
        
        # Store old container number before updating (needed for safe transaction lookup)
        old_container_number = container.container_number
        
        container.container_number = data.get('container_number', container.container_number)
        container.supplier_id = data.get('supplier_id', container.supplier_id)
        container.currency = data.get('currency', container.currency)
        container.exchange_rate = Decimal(str(data.get('exchange_rate', container.exchange_rate)))
        if data.get('date'):
            container.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        container.notes = data.get('notes', container.notes)
        
        # Update expenses - always update if provided
        if 'expense1_amount' in data:
            expense1_amount = data.get('expense1_amount', 0) or 0
            container.expense1_amount = Decimal(str(expense1_amount)) if expense1_amount else None
            # Expense1 always uses container currency and exchange rate
            container.expense1_currency = container.currency  # Use container currency
            container.expense1_exchange_rate = container.exchange_rate  # Use container exchange rate
        
        if 'expense2_amount' in data:
            expense2_amount = data.get('expense2_amount', 0) or 0
            container.expense2_amount = Decimal(str(expense2_amount)) if expense2_amount else None
            expense2_service_id = data.get('expense2_service_company_id')
            container.expense2_service_company_id = int(expense2_service_id) if expense2_service_id else None
            expense2_currency = data.get('expense2_currency')
            container.expense2_currency = expense2_currency if expense2_currency else None
            expense2_exchange_rate = data.get('expense2_exchange_rate')
            container.expense2_exchange_rate = Decimal(str(expense2_exchange_rate)) if expense2_exchange_rate is not None else None
        
        if 'expense3_amount' in data:
            expense3_amount = data.get('expense3_amount', 0) or 0
            old_expense3_base = container.expense3_base_currency
            
            container.expense3_amount = Decimal(str(expense3_amount)) if expense3_amount else None
            expense3_currency = data.get('expense3_currency')
            container.expense3_currency = expense3_currency if expense3_currency else None
            expense3_exchange_rate = data.get('expense3_exchange_rate')
            container.expense3_exchange_rate = Decimal(str(expense3_exchange_rate)) if expense3_exchange_rate is not None else None
            
            # Find safe transaction - try old container number first, then new one
            safe_txn = SafeTransaction.query.filter(
                SafeTransaction.market_id == market_id,
                SafeTransaction.description.like(f'%Container {old_container_number}%Expense 3%')
            ).first()
            
            # If not found with old number, try new number
            if not safe_txn:
                safe_txn = SafeTransaction.query.filter(
                    SafeTransaction.market_id == market_id,
                    SafeTransaction.description.like(f'%Container {container.container_number}%Expense 3%')
                ).first()
            
            if container.expense3_amount and container.expense3_amount > 0:
                new_expense3_base = container.expense3_base_currency
                if safe_txn:
                    # Update existing transaction
                    expense3_rate = container.expense3_exchange_rate or container.exchange_rate
                    exact_base_amount = container.expense3_amount * expense3_rate
                    safe_txn.amount = container.expense3_amount
                    safe_txn.currency = container.expense3_currency or container.currency
                    safe_txn.exchange_rate = expense3_rate
                    safe_txn.amount_base_currency_stored = exact_base_amount  # Store exact value directly
                    safe_txn.date = container.date
                    safe_txn.description = f'Container {container.container_number} - Expense 3 (Cash Expense)'
                else:
                    # Create new transaction - need to recalculate balance after commit
                    # Calculate exact base currency amount
                    expense3_rate = container.expense3_exchange_rate or container.exchange_rate
                    exact_base_amount = container.expense3_amount * expense3_rate
                    
                    safe_txn = SafeTransaction(
                        market_id=market_id,
                        transaction_type='Outflow',
                        amount=container.expense3_amount,
                        currency=container.expense3_currency or container.currency,
                        exchange_rate=expense3_rate,
                        amount_base_currency_stored=exact_base_amount,  # Store exact value directly
                        date=container.date,
                        description=f'Container {container.container_number} - Expense 3 (Cash Expense)',
                        balance_after=Decimal('0')  # Will be recalculated
                    )
                    db.session.add(safe_txn)
            elif safe_txn:
                # Remove safe transaction if expense3 is removed
                db.session.delete(safe_txn)
        
        # Update items if provided
        if 'items' in data:
            # Delete existing items
            PurchaseItem.query.filter_by(container_id=container_id).delete()
            
            # Add new items
            for item_data in data['items']:
                # Ensure item belongs to this supplier
                item_obj = Item.query.get(item_data['item_id'])
                if item_obj:
                    # Update item's supplier_id if not set or different
                    if not item_obj.supplier_id or item_obj.supplier_id != container.supplier_id:
                        item_obj.supplier_id = container.supplier_id
                
                item = PurchaseItem(
                    container_id=container.id,
                    item_id=item_data['item_id'],
                    quantity=Decimal(str(item_data['quantity'])),
                    unit_price=Decimal(str(item_data['unit_price'])),
                    total_price=Decimal(str(item_data['quantity'])) * Decimal(str(item_data['unit_price']))
                )
                db.session.add(item)
        
        db.session.commit()
        
        # Create/update inventory batches if FIFO is active
        from models import Market
        market = Market.query.get(market_id)
        if market and getattr(market, 'calculation_method', 'Average') == 'FIFO':
            from api.fifo_calculations import create_inventory_batches_for_container
            # Delete existing batches and recreate
            from models import InventoryBatch
            InventoryBatch.query.filter_by(container_id=container_id).delete()
            db.session.commit()
            create_inventory_batches_for_container(container_id)
        
        # Recalculate safe balances if expense3 changed
        if 'expense3_amount' in data:
            from api.safe import recalc_safe_balances
            recalc_safe_balances(market_id)
        
        # Refresh container to get updated items
        db.session.refresh(container)
        
        # Return full container with items
        items = [{
            'id': i.id,
            'item_id': i.item_id,
            'item_code': i.item.code,
            'item_name': i.item.name,
            'quantity': float(i.quantity),
            'unit_price': float(i.unit_price),
            'total_price': float(i.total_price)
        } for i in container.items]
        
        return jsonify({
            'id': container.id,
            'container_number': container.container_number,
            'supplier_id': container.supplier_id,
            'supplier_name': container.supplier.name,
            'currency': container.currency,
            'exchange_rate': float(container.exchange_rate),
            'date': container.date.isoformat(),
            'total_amount': float(container.total_amount),
            'notes': container.notes,
            'expense1_amount': float(container.expense1_amount or 0),
            'expense1_currency': container.expense1_currency,
            'expense1_exchange_rate': float(container.expense1_exchange_rate) if container.expense1_exchange_rate is not None else None,
            'expense2_amount': float(container.expense2_amount or 0),
            'expense2_service_company_id': container.expense2_service_company_id,
            'expense2_service_company_name': container.expense_service_company.name if container.expense_service_company else None,
            'expense2_currency': container.expense2_currency,
            'expense2_exchange_rate': float(container.expense2_exchange_rate) if container.expense2_exchange_rate is not None else None,
            'expense3_amount': float(container.expense3_amount or 0),
            'expense3_currency': container.expense3_currency,
            'expense3_exchange_rate': float(container.expense3_exchange_rate) if container.expense3_exchange_rate is not None else None,
            'items': items
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error updating container: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'Error updating container: {str(e)}'}), 500

@bp.route('/containers/<int:container_id>', methods=['DELETE'])
@login_required
def delete_container(container_id):
    market_id = session.get('current_market_id')
    container = PurchaseContainer.query.filter_by(id=container_id, market_id=market_id).first()
    
    if not container:
        return jsonify({'error': 'Container not found'}), 404
    
    # Delete related safe transaction for expense3 if it exists
    safe_txn = SafeTransaction.query.filter(
        SafeTransaction.market_id == market_id,
        SafeTransaction.description.like(f'%Container {container.container_number}%Expense 3%')
    ).first()
    
    if safe_txn:
        db.session.delete(safe_txn)
    
    # Delete the container (purchase items will be deleted via cascade)
    db.session.delete(container)
    db.session.commit()
    
    # Recalculate safe balances after deletion
    from api.safe import recalc_safe_balances
    recalc_safe_balances(market_id)
    
    return jsonify({'success': True})

@bp.route('/by-supplier', methods=['GET'])
@login_required
def get_purchases_by_supplier():
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    supplier_id = request.args.get('supplier_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = PurchaseContainer.query.filter_by(market_id=market_id)
    
    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)
    if start_date:
        query = query.filter(PurchaseContainer.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseContainer.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    containers = query.order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).all()
    
    return jsonify([{
        'id': c.id,
        'container_number': c.container_number,
        'supplier_id': c.supplier_id,
        'supplier_name': c.supplier.name,
        'currency': c.currency,
        'exchange_rate': float(c.exchange_rate),
        'date': c.date.isoformat(),
        'total_amount': float(c.total_amount),
        'notes': c.notes
    } for c in containers])

@bp.route('/by-item', methods=['GET'])
@login_required
def get_purchases_by_item():
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    item_id = request.args.get('item_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = db.session.query(PurchaseItem, PurchaseContainer).join(
        PurchaseContainer, PurchaseItem.container_id == PurchaseContainer.id
    ).filter(PurchaseContainer.market_id == market_id)
    
    if item_id:
        query = query.filter(PurchaseItem.item_id == item_id)
    if start_date:
        query = query.filter(PurchaseContainer.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(PurchaseContainer.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    results = query.order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).all()
    
    return jsonify([{
        'date': container.date.isoformat(),
        'container_number': container.container_number,
        'supplier_name': container.supplier.name,
        'item_code': item.item.code,
        'item_name': item.item.name,
        'quantity': float(item.quantity),
        'unit_price': float(item.unit_price),
        'total_price': float(item.total_price),
        'currency': container.currency,
        'exchange_rate': float(container.exchange_rate)
    } for item, container in results])


@bp.route('/containers/import', methods=['POST'])
@login_required
def import_containers():
    """Import purchase containers and items from Excel.
    Expected columns (case sensitive):
    - ContainerNumber
    - Date (YYYY-MM-DD)
    - Supplier (name)
    - Currency
    - ExchangeRate
    - ItemCode
    - Quantity
    - UnitPrice
    - Notes (optional)
    """
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400

    try:
        try:
            if file.filename.endswith('.xlsx'):
                df = pd.read_excel(file, engine='openpyxl')
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400

        if df.empty:
            return jsonify({'error': 'Excel file is empty'}), 400

        df.columns = df.columns.str.strip()
        required_cols = ['ContainerNumber', 'Date', 'Supplier', 'Currency', 'ExchangeRate', 'ItemCode', 'Quantity', 'UnitPrice']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return jsonify({'error': f'Missing columns: {", ".join(missing)}. Found: {", ".join(df.columns.tolist())}'}), 400

        errors = []
        created_containers = 0
        created_items = 0

        # Cache lookups
        suppliers_by_name = {s.name: s for s in Company.query.filter_by(market_id=market_id, category='Supplier').all()}
        items_by_code = {i.code: i for i in Item.query.filter_by(market_id=market_id).all()}

        # Group rows by container number to create once
        grouped = df.groupby('ContainerNumber')

        for container_number, group in grouped:
            first_row = group.iloc[0]
            supplier_name = str(first_row['Supplier']).strip()
            supplier = suppliers_by_name.get(supplier_name)
            if not supplier:
                errors.append(f'Container {container_number}: Supplier "{supplier_name}" not found')
                continue

            # Parse date
            try:
                date_val = pd.to_datetime(first_row['Date']).date()
            except Exception:
                date_str = str(first_row['Date'])
                errors.append(f'Container {container_number}: Invalid date "{date_str}"')
                continue

            currency = str(first_row['Currency']).strip()
            try:
                exchange_rate = Decimal(str(first_row['ExchangeRate']))
            except Exception:
                errors.append(f'Container {container_number}: Invalid exchange rate')
                continue

            notes = str(first_row['Notes']).strip() if 'Notes' in group.columns and pd.notna(first_row.get('Notes')) else ''

            # Create container
            container = PurchaseContainer(
                market_id=market_id,
                container_number=str(container_number).strip(),
                supplier_id=supplier.id,
                currency=currency,
                exchange_rate=exchange_rate,
                date=date_val,
                notes=notes
            )
            db.session.add(container)
            db.session.flush()
            created_containers += 1

            # Add items
            for idx, row in group.iterrows():
                code = str(row['ItemCode']).strip()
                item = items_by_code.get(code)
                if not item:
                    errors.append(f'Container {container_number}: Item code "{code}" not found (row {idx + 2})')
                    continue

                try:
                    qty = Decimal(str(row['Quantity']))
                    price = Decimal(str(row['UnitPrice']))
                except Exception:
                    errors.append(f'Container {container_number}: Invalid quantity/price for item {code} (row {idx + 2})')
                    continue

                # Ensure item belongs to this supplier
                if not item.supplier_id or item.supplier_id != supplier.id:
                    item.supplier_id = supplier.id

                purchase_item = PurchaseItem(
                    container_id=container.id,
                    item_id=item.id,
                    quantity=qty,
                    unit_price=price,
                    total_price=qty * price
                )
                db.session.add(purchase_item)
                created_items += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'containers_created': created_containers,
            'items_created': created_items,
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Import failed: {str(e)}'}), 400

@bp.route('/export', methods=['GET'])
@login_required
def export_purchases():
    """Export purchases to Excel"""
    market_id = session.get('current_market_id')
    if not market_id:
        return jsonify({'error': 'No market selected'}), 400
    
    supplier_id = request.args.get('supplier_id', type=int)
    
    # Get containers
    query = PurchaseContainer.query.filter_by(market_id=market_id)
    if supplier_id:
        query = query.filter_by(supplier_id=supplier_id)
    
    containers = query.order_by(PurchaseContainer.date.desc(), PurchaseContainer.id.desc()).all()
    
    # Prepare Excel data - one row per container with summary
    export_rows = []
    for container in containers:
        # Get items in this container
        items = PurchaseItem.query.filter_by(container_id=container.id).all()
        
        # Calculate totals
        total_items = len(items)
        total_quantity = sum(float(item.quantity) for item in items)
        total_weight = sum(float(item.item.weight or 0) * float(item.quantity) for item in items if item.item)
        
        # Get expense amounts
        expense1_amount = float(container.expense1_amount) if container.expense1_amount else 0
        expense2_amount = float(container.expense2_amount) if container.expense2_amount else 0
        expense3_amount = float(container.expense3_amount) if container.expense3_amount else 0
        
        # Get service company name for expense2
        expense2_service_company = ''
        if container.expense2_service_company_id:
            service_company = Company.query.get(container.expense2_service_company_id)
            if service_company:
                expense2_service_company = service_company.name
        
        export_rows.append({
            'Container Number': container.container_number,
            'Date': container.date.isoformat() if container.date else '',
            'Supplier': container.supplier.name if container.supplier else '',
            'Currency': container.currency or '',
            'Exchange Rate': float(container.exchange_rate) if container.exchange_rate else 0,
            'Total Amount': float(container.total_amount) if container.total_amount else 0,
            'Total Items': total_items,
            'Total Quantity': total_quantity,
            'Total Weight': total_weight,
            'Expense 1 Amount': expense1_amount,
            'Expense 2 Service Company': expense2_service_company,
            'Expense 2 Amount': expense2_amount,
            'Expense 3 Amount': expense3_amount,
            'Total Expenses': expense1_amount + expense2_amount + expense3_amount,
            'Notes': container.notes or ''
        })
    
    # Create Excel file
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if export_rows:
            # Summary sheet
            df_summary = pd.DataFrame(export_rows)
            df_summary.to_excel(writer, index=False, sheet_name='Containers Summary')
            
            # Format summary sheet
            from openpyxl.utils import get_column_letter
            worksheet_summary = writer.sheets['Containers Summary']
            for idx, col in enumerate(df_summary.columns):
                max_length = max(
                    df_summary[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                worksheet_summary.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
            
            # Detailed items sheet
            items_rows = []
            for container in containers:
                items = PurchaseItem.query.filter_by(container_id=container.id).all()
                for item in items:
                    items_rows.append({
                        'Container Number': container.container_number,
                        'Date': container.date.isoformat() if container.date else '',
                        'Supplier': container.supplier.name if container.supplier else '',
                        'Item Code': item.item.code if item.item else '',
                        'Item Name': item.item.name if item.item else '',
                        'Quantity': float(item.quantity),
                        'Unit Price': float(item.unit_price),
                        'Total Price': float(item.total_price),
                        'Unit Weight': float(item.item.weight) if item.item and item.item.weight else 0,
                        'Total Weight': float(item.item.weight or 0) * float(item.quantity) if item.item else 0,
                        'Currency': container.currency or '',
                        'Exchange Rate': float(container.exchange_rate) if container.exchange_rate else 0
                    })
            
            if items_rows:
                df_items = pd.DataFrame(items_rows)
                df_items.to_excel(writer, index=False, sheet_name='Items Detail')
                
                # Format items sheet
                worksheet_items = writer.sheets['Items Detail']
                for idx, col in enumerate(df_items.columns):
                    max_length = max(
                        df_items[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet_items.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
    
    output.seek(0)
    filename = f'purchases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name=filename)
